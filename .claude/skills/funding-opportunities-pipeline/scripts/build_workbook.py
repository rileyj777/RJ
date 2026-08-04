#!/usr/bin/env python3
# SKILL NOTE (funding-opportunities-pipeline, Workflow 5):
# This is a STABLE SCAFFOLD. It bakes in NO account/contact/angle/gap data and
# NO scoring taxonomy. Everything comes from disk at run time:
#   - account/contact/angle data  -> JSON files you produce from LIVE monday
#     pulls (Workflows 2 & 3) right before the build (schemas below)
#   - scoring taxonomy (kw tiers, junk filter, capability tags, ICP weights)
#     -> scoring_taxonomy.json (canonical; shipped next to this script)
# Nothing here is a dated snapshot, so there are no frozen blocks to hand-edit
# and no hardcoded upload filenames to crash on. Counts and the zero-contact /
# no-economic-buyer red-flag lists are COMPUTED from the data at build time.
#
# Run (from the dir holding your JSON inputs):
#   python build_workbook.py --book book.json --contacts contacts.json \
#       --angle angle.json --gap gap_candidates.json
# Every input except --book is optional; a missing optional file is skipped
# with a warning (no crash). --taxonomy defaults to the JSON beside this file.
# After building, ALWAYS run:
#   python /mnt/skills/public/xlsx/scripts/recalc.py <output.xlsx>
# and ship only on status=success with zero errors.
#
# ---------------------------------------------------------------------------
# INPUT SCHEMAS (all JSON, UTF-8)
#
# book.json  (REQUIRED) - your book from the WF2 pull, one object per account:
#   [{"name": "Cool Amps", "status": "Prospect",
#     "loc": "Golden, CO", "ind": "Energy",
#     "fit": "Battery, Recycling", "id": "12345"}, ...]
#   name + status required; loc/ind/fit/id optional. status in
#   Suspect|Prospect|Lead (others sort last).
#
# contacts.json  (optional) - from the WF3 pull, one object per contact:
#   [{"account": "Cool Amps", "name": "Lonnie Garris", "title": "CEO",
#     "seniority": "C-Suite", "email": "lonnie@coolamps.tech",
#     "role": "Economic Buyer", "flag": ""}, ...]
#   role in Economic Buyer|Technical Buyer|User Buyer|"" (blank = untagged;
#   never guess a role - a blank beats a wrong tag). flag = a data-quality
#   warning string (e.g. "wrong company", "verify email"); flagged rows render
#   pink. Coaches are never machine-assigned.
#
# angle.json  (optional) - funding angle per account, read from the pipeline
#   board's Related-accounts / your mapping. {"Cool Amps": "Argonne ReCell
#   license + CMI Li-ion leach patent", ...}. Matched to accounts by
#   normalized name; loose (substring) matches are reported for your review.
#
# gap_candidates.json  (optional) - normalized gap-list companies (WF4 output
#   from whatever lists were uploaded THIS session), one object per (company,
#   list) appearance; the script merges by name, scores with the taxonomy,
#   drops junk / in-book / score<3, and flags multi-list hits:
#   [{"name": "Foo Corp", "source": "ARPA-E Summit 2026",
#     "description": "battery recycling pilot", "prequalified": false}, ...]
#   If absent, the Gap Analysis tab is skipped entirely (not emitted empty).
# ---------------------------------------------------------------------------
"""Build the JE Strategic Selling Workbook: live ICP scoring, contact map,
blue sheet, gap analysis. Scaffold only - data and taxonomy come from disk."""
import argparse, json, os, re, difflib, unicodedata, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from functools import lru_cache
from collections import defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def load_json(path, required=False, default=None, label=''):
    """Read a JSON file; missing optional files return `default` with a warning."""
    if not path or not os.path.exists(path):
        if required:
            sys.exit(f"ERROR: required input {label or path} not found at {path}")
        print(f"  (skip) {label or path}: not provided - continuing without it")
        return default
    with open(path, encoding='utf-8') as f:
        return json.load(f)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--book', default='book.json', help='REQUIRED account data (WF2 pull)')
    ap.add_argument('--contacts', default='contacts.json', help='contact data (WF3 pull)')
    ap.add_argument('--angle', default='angle.json', help='funding-angle map')
    ap.add_argument('--gap', default='gap_candidates.json', help='normalized gap-list candidates (WF4)')
    ap.add_argument('--taxonomy', default=os.path.join(SCRIPT_DIR, 'scoring_taxonomy.json'),
                    help='canonical scoring taxonomy (defaults beside this script)')
    ap.add_argument('--out', default='JE_Strategic_Selling_Workbook.xlsx', help='output .xlsx path')
    args = ap.parse_args()

    print("Loading inputs...")
    TAX = load_json(args.taxonomy, required=True, label='scoring_taxonomy.json')
    book_list = load_json(args.book, required=True, label='book.json')
    contacts_raw = load_json(args.contacts, default=[], label='contacts.json')
    angle = load_json(args.angle, default={}, label='angle.json')
    gap_raw = load_json(args.gap, default=None, label='gap_candidates.json')

    # ---------- taxonomy-derived matchers ----------
    KW = {int(k): v for k, v in TAX['keyword_tiers'].items()}
    # compile the junk regex once
    JUNK = re.compile(r'\b(' + '|'.join(TAX['junk_terms']) + r')\b', re.I)
    STRONG = [t.lower() for t in TAX['bullseye_tags']]
    ADJACENT = [t.lower() for t in TAX['adjacent_tags']]
    CORE = STRONG + ADJACENT
    VALID_IND = set(TAX['industry_fit_valid_industries'])
    WEIGHTS_MAP = TAX['icp_weights']
    FUZZY = float(TAX.get('fuzzy_cutoff', 0.92))
    INCLUDE = int(TAX.get('score_include_threshold', 3))
    PREQUAL = int(TAX.get('prequalified_force_include', 4))
    SUFFIX = re.compile(r'\b(' + '|'.join(re.escape(s) for s in TAX['normalize_suffix_terms']) + r')\b\.?', re.I)

    # Precompile combined regexes for keyword tiers to avoid nested Python loops
    KW_RE = {}
    for sc, lst in KW.items():
        # join escaped terms so literal substrings are matched as before
        try:
            pattern = '|'.join(re.escape(k) for k in lst)
            KW_RE[sc] = re.compile(pattern, re.I)
        except Exception:
            # fallback: compile each separately (defensive)
            KW_RE[sc] = [re.compile(re.escape(k), re.I) for k in lst]

    @lru_cache(maxsize=4096)
    def norm(s):
        s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode()
        s = re.sub(r'\([^)]*\)', ' ', s)
        s = re.sub(r'[^a-z0-9 ]', ' ', s.lower())
        s = SUFFIX.sub(' ', s)
        return re.sub(r'\s+', ' ', s).strip()

    @lru_cache(maxsize=8192)
    def industry_fit(fit, ind):
        t = (fit or '').lower()
        strong = sum(1 for k in STRONG if k in t)
        total = sum(1 for k in CORE if k in t)
        if strong >= 2 or total >= 3: return 5
        if strong == 1 or total == 2: return 4
        if total == 1: return 3
        if (ind or '') in VALID_IND: return 2
        if ind: return 1
        return ''

    @lru_cache(maxsize=4096)
    def kw_score(text):
        t = ' ' + re.sub(r'[^a-z0-9 ]', ' ', str(text).lower()) + ' '
        # try compiled combined regex first
        for sc in (5, 4, 3):
            cre = KW_RE.get(sc)
            if cre is None:
                continue
            if isinstance(cre, list):
                for rx in cre:
                    m = rx.search(t)
                    if m: return sc, m.group(0).strip()
            else:
                m = cre.search(t)
                if m: return sc, m.group(0).strip()
        return 0, ''

    # ---------- book (dict keyed by name; preserves input order) ----------
    book = {}
    for e in book_list:
        nm = str(e.get('name', '')).strip()
        if not nm: continue
        book[nm] = dict(id=e.get('id', ''), status=e.get('status', ''),
                        loc=e.get('loc', ''), ind=e.get('ind', ''), fit=e.get('fit', ''))
    # build normalized map once
    book_norm = {norm(n): n for n in book}
    book_norm_keys = list(book_norm.keys())

    # Build a simple prefix index to reduce candidates for fuzzy matching
    prefix_index = defaultdict(list)
    for key in book_norm_keys:
        # use prefixes of lengths 3..6 to bucket similar names
        for L in range(3, min(7, len(key) + 1)):
            prefix_index[key[:L]].append(key)

    print(f"Book: {len(book)} accounts")

    def in_book(name):
        n = norm(name)
        if not n: return None
        if n in book_norm: return book_norm[n]
        # fast substring/starts-with checks first (cheap)
        for bk in book_norm_keys:
            if bk and n and (bk == n or (len(n) > 4 and len(bk) > 4 and (bk.startswith(n) or n.startswith(bk)))):
                return book_norm[bk]
        # build candidate set from prefix buckets (try longer prefixes first)
        candidates = set()
        for L in range(min(6, len(n)), 2, -1):
            candidates.update(prefix_index.get(n[:L], []))
            if candidates:
                break
        if candidates:
            close = difflib.get_close_matches(n, list(candidates), n=1, cutoff=FUZZY)
            return book_norm[close[0]] if close else None
        # final fallback: no good prefix bucket, give up
        return None

    # ---------- contacts ----------
    E, T, U = 'Economic Buyer', 'Technical Buyer', 'User Buyer'
    contacts = []
    for c in contacts_raw:
        role = str(c.get('role', '') or '').strip()
        flag = str(c.get('flag', '') or '').strip()
        if not flag and role.startswith('⚠'):  # legacy: warning carried in role
            flag, role = role, ''
        contacts.append(dict(account=str(c.get('account', '')).strip(), name=c.get('name', ''),
                             title=c.get('title', ''), seniority=c.get('seniority', ''),
                             email=c.get('email', ''), role=role if role in (E, T, U) else '', flag=flag))
    print(f"Contacts: {len(contacts)}")

    # ---------- red flags (COMPUTED, not frozen) ----------
    # Only meaningful when contacts were actually loaded; scoped to accounts in
    # play (Prospect/Lead) so unengaged Suspects don't flood the list.
    red_zero, red_no_econ = [], []
    if contacts:
        by_acct = {}
        for c in contacts:
            by_acct.setdefault(norm(c['account']), []).append(c)
        for nm, d in book.items():
            if d['status'] not in ('Prospect', 'Lead'): continue
            cs = by_acct.get(norm(nm), [])
            if not cs:
                red_zero.append(nm)
            elif not any(c['role'] == E for c in cs):
                red_no_econ.append(nm)
        red_zero.sort(); red_no_econ.sort()

    # ---------- gap analysis (COMPUTED from normalized candidates) ----------
    gaps, multi = [], []
    if gap_raw:
        gap = {}
        for cand in gap_raw:
            name = str(cand.get('name', '')).strip()
            if not name or len(name) < 3 or name.lower() == 'nan': continue
            if JUNK.search(name): continue
            n = norm(name)
            if not n or in_book(name): continue
            source = str(cand.get('source', 'list')).strip() or 'list'
            desc = str(cand.get('description', '') or '')[:160]
            prequal = bool(cand.get('prequalified', False))
            sc, kw = kw_score(name + ' ' + desc)
            if prequal:
                sc, kw = max(sc, PREQUAL), (kw or 'pre-qualified')
            if sc < INCLUDE: continue
            if n in gap:
                gap[n]['sources'].add(source)
                if sc > gap[n]['score']: gap[n]['score'], gap[n]['kw'] = sc, kw
                if desc and len(desc) > len(gap[n]['desc']): gap[n]['desc'] = desc
            else:
                gap[n] = dict(name=name, sources={source}, desc=desc, score=sc, kw=kw)
        gaps = sorted(gap.values(), key=lambda d: (-d['score'], -len(d['sources']), d['name']))
        multi = [g for g in gaps if len(g['sources']) > 1]
        print(f"Gap candidates (qualified, not in book): {len(gaps)}; on multiple lists: {len(multi)}")
    else:
        print("  (skip) Gap Analysis tab: no gap_candidates.json this run")

    # ================= BUILD WORKBOOK =================
    wb = openpyxl.Workbook()
    AR = 'Arial'
    def Fn(sz=10, b=False, c='000000'): return Font(name=AR, size=sz, bold=b, color=c)
    BLUE = Font(name=AR, size=10, color='0000FF'); YEL = PatternFill('solid', fgColor='FFFF00')
    HDR = PatternFill('solid', fgColor='1F3864'); HDRF = Font(name=AR, size=10, bold=True, color='FFFFFF')
    SUB = PatternFill('solid', fgColor='D9E1F2'); GOLD = PatternFill('solid', fgColor='FFC000')
    thin = Side(style='thin', color='BFBFBF'); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_acct = len(book)
    n_contact = len(contacts)
    n_pretag = sum(1 for c in contacts if c['role'] in (E, T, U))

    # ---- READ ME ----
    ws = wb.active; ws.title = 'READ ME'
    ws.column_dimensions['A'].width = 118
    contact_line = (f"Built from the 2026 Strategic Selling training (Miller/Heiman) + your monday CRM "
                    f"({n_acct} accounts, {n_contact} contacts at top targets.)")
    rows = [
        ("JOHNSTON ENGINEERING - STRATEGIC SELLING WORKBOOK", 14, True),
        (contact_line, 10, False),
        ("", 10, False),
        ("HOW THE LIVE SCORING WORKS (Account Scorecard tab)", 11, True),
        ("- The 5 Ideal Customer Criteria from the training: Industry Fit, Company Size Fit, Compelling Event Present, Access to Economic Buyer, Growth/Follow-On Work Potential. Score each 0-5 (5...", 10, False),
        ("- Row 2 holds the WEIGHTS (blue cells). Change a weight and every account re-scores and re-ranks instantly. Defaults favor what opens JE deals: Compelling Event Present = 2.0, Access to Economic Buyer = 1.5, rest 1.0", 10, False),
        ("- Weighted Score % = SUMPRODUCT(your scores x weights) / max possible. Rank and A-D tier update automatically.", 10, False),
        ("- Industry Fit is pre-scored from your CRM capability tags (blue = my suggestion, edit freely). Yellow cells are yours to fill - blanks count as 0, and the 'Criteria left' column shows how many you haven't filled.", 10, False),
        ("", 10, False),
        ("COLOR LEGEND: blue text = editable input/suggestion - yellow fill = fill these in - black = live formula (don't overwrite)", 10, True),
        ("", 10, False),
        ("TABS", 11, True),
        (f"- Account Scorecard - all {n_acct} accounts, live weighted ICP ranking + funding angle from your pipeline board.", 10, False),
        (f"- Buying Influences - contact map for top accounts. Role/Degree/Mode/Rating dropdowns straight from the training. {n_pretag} obvious Economic/Technical/User buyers are pre-tagged; Coach never auto-assigned.", 10, False),
        ("- Blue Sheet - one-deal Strategic Analysis Plan modeled on your JE template. Duplicate the tab per deal.", 10, False),
    ]
    # Data-quality note is dynamic: list whatever contact flags actually loaded.
    flagged = [c for c in contacts if c['flag']]
    if flagged:
        rows.append(("", 10, False))
        note = "Data note: " + "; ".join(f"{c['account']} - {c['name']}: {c['flag']}" for c in flagged[:8])
        if len(flagged) > 8: note += f"; (+{len(flagged) - 8} more flagged in Buying Influences)"
        rows.append((note, 10, False))
    for i, (t, sz, b) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=t).font = Fn(sz, b)
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical='top')

    # ---- Lists (validation sources) ----
    wl = wb.create_sheet('Lists')
    lists = {
        'A': ('Role', ['Economic Buyer', 'User Buyer', 'Technical Buyer', 'Coach']),
        'B': ('Degree', ['High', 'Medium', 'Low']),
        'C': ('Mode', ['Growth', 'Trouble', 'Even Keel', 'Overconfident']),
        'D': ('Rating', ['+5 Strong Positive', '+3 Mild Positive', '0 Neutral', '-3 Mild Negative', '-5 Strong Negative']),
        'E': ('Competition', ['Exclusive', 'Dominant', 'Shared', 'Zero']),
        'F': ('Funnel', ['Universe', 'Above', 'In', 'Best Few', 'Won', 'Lost']),
        'G': ('Timing', ['Urgent', 'Active', 'Work It In', 'Later']),
        'H': ('Score05', ['0', '1', '2', '3', '4', '5']),
    }
    for col, (nm, vals) in lists.items():
        wl[f'{col}1'] = nm; wl[f'{col}1'].font = Fn(10, True)
        for i, v in enumerate(vals, 2): wl[f'{col}{i}'] = v; wl[f'{col}{i}'].font = Fn()
    wl.sheet_state = 'hidden'

    # ---- Account Scorecard ----
    sc = wb.create_sheet('Account Scorecard')
    crit = ['Industry Fit', 'Company Size Fit', 'Compelling Event Present', 'Access to Economic Buyer', 'Growth/Follow-On Potential']
    WEIGHTS = [WEIGHTS_MAP[c] for c in crit]
    sc['A1'] = 'ACCOUNT SCORECARD - live weighted ICP ranking (edit blue/yellow; formulas do the rest)'; sc['A1'].font = Fn(12, True)
    sc['A2'] = 'WEIGHTS ->'; sc['A2'].font = Fn(10, True); sc['A2'].alignment = Alignment(horizontal='right')
    headers = ['Account', 'Status', 'Industry', 'Capability tags', 'Location', 'Funding angle (from pipeline board)'] + crit + ['Weighted Score %', 'Rank', 'Tier', 'Criteria left']
    CI = {h: i + 1 for i, h in enumerate(headers)}
    for h, i in CI.items():
        c = sc.cell(row=3, column=i, value=h); c.font = HDRF; c.fill = HDR; c.alignment = Alignment(wrap_text=True, vertical='center'); c.border = BORD
    for j in range(len(crit)):
        w = sc.cell(row=2, column=CI['Industry Fit'] + j, value=WEIGHTS[j]); w.font = BLUE; w.fill = YEL; w.border = BORD
    WT1 = get_column_letter(CI['Industry Fit']); WT2 = get_column_letter(CI['Growth/Follow-On Potential'])
    acct_rows = sorted(book.items(), key=lambda kv: ({'Lead': 0, 'Prospect': 1, 'Suspect': 2}.get(kv[1]['status'], 3), kv[0].lower()))

    # angle match with review reporting
    angle_norm = {norm(k): (k, v) for k, v in angle.items()}
    matched_angle_keys, loose_matches = set(), []
    def angle_for(name):
        nn = norm(name)
        if nn in angle_norm:
            matched_angle_keys.add(angle_norm[nn][0]); return angle_norm[nn][1]
        for kn, (k, v) in angle_norm.items():
            if kn and (kn in nn or nn in kn):
                matched_angle_keys.add(k); loose_matches.append((k, name)); return v
        return ''

    r = 4
    for name, d in acct_rows:
        sc.cell(row=r, column=CI['Account'], value=name).font = Fn(10, True)
        sc.cell(row=r, column=CI['Status'], value=d['status']).font = Fn()
        sc.cell(row=r, column=CI['Industry'], value=d['ind']).font = Fn()
        sc.cell(row=r, column=CI['Capability tags'], value=d['fit']).font = Fn(9)
        sc.cell(row=r, column=CI['Location'], value=d['loc']).font = Fn(9)
        sc.cell(row=r, column=CI['Funding angle (from pipeline board)'], value=angle_for(name)).font = Fn(9, c='7030A0')
        ifit = industry_fit(d['fit'], d['ind'])
        c1 = sc.cell(row=r, column=CI['Industry Fit'], value=ifit if ifit != '' else None); c1.font = BLUE; c1.border = BORD
        for j in range(1, 5):
            cc = sc.cell(row=r, column=CI['Industry Fit'] + j); cc.fill = YEL; cc.border = BORD; cc.font = BLUE
        sr1 = f'{WT1}{r}'; sr2 = f'{WT2}{r}'
        sc.cell(row=r, column=CI['Weighted Score %'], value=f'=IFERROR(SUMPRODUCT({sr1}:{sr2},${WT1}$2:${WT2}$2)/(5*SUM(${WT1}$2:${WT2}$2)),0)').font = Fn(10, True)
        sc.cell(row=r, column=CI['Weighted Score %']).number_format = '0.0%'
        last = 3 + len(acct_rows)
        WS = get_column_letter(CI['Weighted Score %'])
        sc.cell(row=r, column=CI['Rank'], value=f'=IF({WS}{r}=0,"",RANK({WS}{r},${WS}$4:${WS}${last}))').font = Fn()
        sc.cell(row=r, column=CI['Tier'], value=f'=IF({WS}{r}>=0.75,"A",IF({WS}{r}>=0.55,"B",IF({WS}{r}>=0.35,"C",IF({WS}{r}>0,"D","-"))))').font = Fn(10, True)
        sc.cell(row=r, column=CI['Criteria left'], value=f'=COUNTBLANK({sr1}:{sr2})').font = Fn(9, c='808080')
        r += 1
    widths = {'Account': 30, 'Status': 9, 'Industry': 17, 'Capability tags': 38, 'Location': 22, 'Funding angle (from pipeline board)': 40, 'Weighted Score %': 11, 'Rank': 6, 'Tier': 6, 'Criteria[...]': 10}
    for h, wd in widths.items(): sc.column_dimensions[get_column_letter(CI[h])].width = wd
    for cname in crit: sc.column_dimensions[get_column_letter(CI[cname])].width = 10
    sc.freeze_panes = 'B4'
    dv05 = DataValidation(type='list', formula1='=Lists!$H$2:$H$7', allow_blank=True)
    sc.add_data_validation(dv05); dv05.add(f'{WT1}4:{WT2}{3 + len(acct_rows)}')

    # ---- Buying Influences ----
    bi = wb.create_sheet('Buying Influences')
    bi['A1'] = "BUYING INFLUENCES - contact map (Roles: E=final approval - U=lives with the work - T=screens on specs - C=coach, develop don't assign)"; bi['A1'].font = Fn(12, True)
    bh = ['Account', 'Contact', 'Title', 'Seniority', 'Email', 'Role (E/U/T/C)', 'Degree (H/M/L)', 'Mode', 'Rating', 'Win-Results statement (their personal win)', 'Base covered? evidence / notes']
    for i, h in enumerate(bh, 1):
        c = bi.cell(row=2, column=i, value=h); c.font = HDRF; c.fill = HDR; c.alignment = Alignment(wrap_text=True); c.border = BORD
    r = 3
    for c in contacts:
        vals = [c['account'], c['name'], c['title'], c['seniority'], c['email']]
        for i, v in enumerate(vals, 1): bi.cell(row=r, column=i, value=v).font = Fn(9)
        rc = bi.cell(row=r, column=6, value=c['role'] if c['role'] in (E, T, U) else None)
        rc.font = BLUE; rc.border = BORD
        if c['flag']:
            bi.cell(row=r, column=11, value=c['flag']).font = Fn(9, True, 'C00000')
            for i in range(1, 12): bi.cell(row=r, column=i).fill = PatternFill('solid', fgColor='FCE4EC')
        for col in (7, 8, 9):
            cc = bi.cell(row=r, column=col); cc.fill = YEL; cc.border = BORD; cc.font = BLUE
        bi.cell(row=r, column=10).fill = YEL
        r += 1
    if contacts:
        r += 1
        if red_zero:
            bi.cell(row=r, column=1, value='RED FLAGS - Prospect/Lead accounts with ZERO contacts in CRM (uncovered bases per the training):').font = Fn(10, True, 'C00000'); r += 1
            for nm in red_zero:
                bi.cell(row=r, column=1, value=nm).font = Fn(10, c='C00000')
                bi.cell(row=r, column=2, value='No buying influences identified - critical information missing').font = Fn(9, c='C00000')
                r += 1
        if red_no_econ:
            r += 1
            bi.cell(row=r, column=1, value='RED FLAGS - accounts with contacts but NO Economic Buyer (no known line to the money):').font = Fn(10, True, 'C00000'); r += 1
            for nm in red_no_econ:
                bi.cell(row=r, column=1, value=nm).font = Fn(10, c='C00000')
                bi.cell(row=r, column=2, value='Has contacts, none tagged Economic Buyer - find the approver').font = Fn(9, c='C00000')
                r += 1
    for col, wd in zip('ABCDEFGHIJK', [24, 22, 34, 17, 30, 15, 12, 13, 16, 42, 32]): bi.column_dimensions[col].width = wd
    bi.freeze_panes = 'A3'
    if contacts:
        for col, f1 in ((6, '=Lists!$A$2:$A$5'), (7, '=Lists!$B$2:$B$4'), (8, '=Lists!$C$2:$C$5'), (9, '=Lists!$D$2:$D$6')):
            dv = DataValidation(type='list', formula1=f1, allow_blank=True); bi.add_data_validation(dv)
            dv.add(f'{get_column_letter(col)}3:{get_column_letter(col)}{2 + len(contacts)}')

    # ---- Blue Sheet ----
    bs = wb.create_sheet('Blue Sheet')
    bs['A1'] = 'STRATEGIC ANALYSIS PLAN (Blue Sheet) - Johnston Engineering | Confidential & Proprietary'; bs['A1'].font = Fn(12, True)
    bs['A2'] = 'Duplicate this tab per sales objective. Objective format: TO SELL {service} to {company/area} FOR {revenue} BY {close date}.'; bs['A2'].font = Fn(9)
    left = [('BDR / Owner', ''), ('Customer Name', ''), ('Sales Objective', ''), ('Potential Revenue', ''), ('Close Date', ''), ('Initial Start Date', ''), ('Last Updated', '=TODAY()')]
    r = 4
    for lab, val in left:
        bs.cell(row=r, column=1, value=lab).font = Fn(10, True)
        c = bs.cell(row=r, column=2, value=val if val else None)
        c.fill = YEL if not val else PatternFill(); c.font = BLUE if not val else Fn(); c.border = BORD
        if lab in ('Close Date', 'Initial Start Date'): c.number_format = 'yyyy-mm-dd'
        if lab == 'Last Updated': c.number_format = 'yyyy-mm-dd'; c.font = Fn()
        r += 1
    pos = [('My Position VS Competition', '=Lists!$E$2:$E$5'), ('Place in Sales Funnel', '=Lists!$F$2:$F$7'), ('Timing for Priorities', '=Lists!$G$2:$G$5')]
    r = 4
    for lab, f1 in pos:
        bs.cell(row=r, column=4, value=lab).font = Fn(10, True)
        c = bs.cell(row=r, column=6); c.fill = YEL; c.border = BORD; c.font = BLUE
        dv = DataValidation(type='list', formula1=f1, allow_blank=True); bs.add_data_validation(dv); dv.add(c.coordinate)
        r += 1
    bs.cell(row=7, column=4, value='"How do I feel right now about closing this?" (-5 PANIC ... +5 EUPHORIA)').font = Fn(9)
    cfeel = bs.cell(row=7, column=6); cfeel.fill = YEL; cfeel.border = BORD; cfeel.font = BLUE
    bs.cell(row=4, column=8, value='IDEAL CUSTOMER CRITERIA (5 = best aligned, 0 = worst)').font = Fn(10, True)
    for i, cn in enumerate(crit):
        bs.cell(row=5 + i, column=8, value=cn).font = Fn(10)
        cc = bs.cell(row=5 + i, column=10); cc.fill = YEL; cc.border = BORD; cc.font = BLUE
        dv = DataValidation(type='list', formula1='=Lists!$H$2:$H$7', allow_blank=True); bs.add_data_validation(dv); dv.add(cc.coordinate)
    bs.cell(row=10, column=8, value='TOTAL (max 25)').font = Fn(10, True)
    bs.cell(row=10, column=10, value='=SUM(J5:J9)').font = Fn(10, True)
    gh = ['Buying Influence (Name, Title, Location)', 'Role', 'Degree', 'Mode', 'Rating', 'Key Win-Result (personal win statement)', 'How well is base covered? Evidence:']
    for i, h in enumerate(gh, 1):
        c = bs.cell(row=13, column=i, value=h); c.font = HDRF; c.fill = HDR; c.alignment = Alignment(wrap_text=True); c.border = BORD
    for rr in range(14, 24):
        for i in range(1, 8):
            c = bs.cell(row=rr, column=i); c.border = BORD
            if i in (2, 3, 4, 5): c.fill = YEL
    for col, f1 in ((2, '=Lists!$A$2:$A$5'), (3, '=Lists!$B$2:$B$4'), (4, '=Lists!$C$2:$C$5'), (5, '=Lists!$D$2:$D$6')):
        dv = DataValidation(type='list', formula1=f1, allow_blank=True); bs.add_data_validation(dv)
        dv.add(f'{get_column_letter(col)}14:{get_column_letter(col)}23')
    sect = [(25, 'STRENGTHS (differentiation; relevant to objective; diminish price)', '00B050'), (31, 'RED FLAGS (missing info - untested assumption - uncovered base - new player - reorg)', 'C00[...])']
    for row0, lab, color in sect:
        bs.cell(row=row0, column=1, value=lab).font = Fn(10, True, color)
        for rr in range(row0 + 1, row0 + 5):
            bs.cell(row=rr, column=1).border = BORD
            bs.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5)
    bs.cell(row=25, column=6, value='BEST ACTION PLAN').font = Fn(10, True)
    for i, h in enumerate(['What', 'Who', 'When'], 6): bs.cell(row=26, column=i, value=h).font = HDRF; bs.cell(row=26, column=i).fill = HDR
    for rr in range(27, 32):
        for i in (6, 7, 8): bs.cell(row=rr, column=i).border = BORD
    bs.cell(row=33, column=6, value='INFORMATION NEEDED / FROM WHOM').font = Fn(10, True)
    for rr in range(34, 37):
        for i in (6, 7): bs.cell(row=rr, column=i).border = BORD
    for col, wd in zip('ABCDEFGHIJ', [34, 14, 10, 12, 16, 26, 26, 30, 4, 8]): bs.column_dimensions[col].width = wd

    # ---- Gap Analysis (only if candidates were provided) ----
    if gaps:
        ga = wb.create_sheet('Gap Analysis')
        ga['A1'] = f'GAP ANALYSIS - qualified companies from your lists NOT in your {n_acct}-account book (sorted by fit; * = on 2+ lists)'
        ga['A1'].font = Fn(12, True)
        gh2 = ['Company', 'Fit (0-5)', 'On lists', 'Source list(s)', 'Evidence / description', 'Keyword hit', 'Suggested next step']
        for i, h in enumerate(gh2, 1):
            c = ga.cell(row=2, column=i, value=h); c.font = HDRF; c.fill = HDR; c.border = BORD
        r = 3
        for g in gaps:
            star = '* ' if len(g['sources']) > 1 else ''
            ga.cell(row=r, column=1, value=star + g['name']).font = Fn(10, len(g['sources']) > 1)
            ga.cell(row=r, column=2, value=g['score']).font = Fn()
            ga.cell(row=r, column=3, value=len(g['sources'])).font = Fn()
            ga.cell(row=r, column=4, value=', '.join(sorted(g['sources']))).font = Fn(9)
            ga.cell(row=r, column=5, value=g['desc']).font = Fn(9)
            ga.cell(row=r, column=6, value=g['kw']).font = Fn(9, c='7030A0')
            step = 'Add as Suspect + research' if g['score'] >= 4 else 'Quick qualify vs ICP'
            if len(g['sources']) > 1: step = 'PRIORITY: add as Suspect - multi-list signal'
            ga.cell(row=r, column=7, value=step).font = Fn(9)
            if len(g['sources']) > 1:
                for i in range(1, 8): ga.cell(row=r, column=i).fill = PatternFill('solid', fgColor='FFF2CC')
            r += 1
        for col, wd in zip('ABCDEFG', [34, 8, 8, 34, 60, 18, 34]): ga.column_dimensions[col].width = wd
        ga.freeze_panes = 'A3'
        ga.auto_filter.ref = f'A2:G{r - 1}'

    # tab order & colors
    wb.move_sheet('READ ME', offset=0)
    tabcolors = [('READ ME', '1F3864'), ('Account Scorecard', '00B050'), ('Buying Influences', 'FFC000'), ('Blue Sheet', '7030A0'), ('Gap Analysis', 'C00000')]
    for nm, c in tabcolors:
        if nm in wb.sheetnames: wb[nm].sheet_properties.tabColor = c

    wb.save(args.out)
    print('\nSaved', args.out)
    print(f"Scorecard rows: {len(acct_rows)} | Contacts: {n_contact} (pre-tagged: {n_pretag}) | "
          f"Gap rows: {len(gaps)} (multi-list: {len(multi)})")
    if contacts:
        print(f"Red flags: {len(red_zero)} zero-contact, {len(red_no_econ)} no-economic-buyer (Prospect/Lead)")

    # ---- coverage / review warnings (skill ethos: declare what needs eyeballing) ----
    unmatched = [k for k in angle if k not in matched_angle_keys]
    if unmatched:
        print(f"\nWARNING: {len(unmatched)} funding-angle entries matched NO account (dropped, verify names): {', '.join(unmatched)}")
    if loose_matches:
        print(f"WARNING: {len(loose_matches)} angle(s) matched by loose substring - confirm they landed right:")
        for k, nm in loose_matches: print(f"    angle '{k}' -> account '{nm}'")
    if gaps:
        print("Top 15 gap candidates:")
        for g in gaps[:15]:
            print(f"  [{g['score']}] {'*' if len(g['sources']) > 1 else ' '} {g['name']} <- {', '.join(sorted(g['sources']))} | {g['desc'][:60]}")


if __name__ == '__main__':
    main()
