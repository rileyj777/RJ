"""End-to-end smoke test for build_workbook.py.

The script is a single large main() with nested closures, so there is nothing
importable to unit-test. Instead this runs it the way Workflow 5 runs it - as a
subprocess over JSON fixtures - and asserts the workbook comes out with the
sheets and the scoring the skill documents.

This is the test that would have caught the 2026-08 re.escape regression: it
asserts a solid-state battery company lands at fit 5, not 4.
"""
import json
import subprocess
import sys
from pathlib import Path

import pytest

# Hard import, deliberately not importorskip: openpyxl is declared in
# requirements.txt and CI installs it, so a missing openpyxl means the
# environment is wrong. Skipping here would silently hide the only
# end-to-end test and leave CI green on a broken builder.
import openpyxl  # noqa: E402  (install with: pip install -r requirements.txt)

REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = (
    REPO_ROOT / ".claude/skills/funding-opportunities-pipeline/scripts/build_workbook.py"
)

BOOK = [
    {"name": "Cool Amps", "status": "Prospect", "loc": "Golden, CO",
     "ind": "Energy", "fit": "Battery, Recycling", "id": "1001"},
    {"name": "Hydrolyte Works", "status": "Lead", "loc": "Spokane, WA",
     "ind": "Energy", "fit": "Electrolyzer, Hydrogen", "id": "1002"},
    {"name": "Quiet Cases", "status": "Suspect", "loc": "Boise, ID",
     "ind": "Manufacturing", "fit": "", "id": "1003"},
]

CONTACTS = [
    {"account": "Cool Amps", "name": "Lonnie Garris", "title": "CEO",
     "seniority": "C-Suite", "email": "lonnie@coolamps.tech",
     "role": "Economic Buyer", "flag": ""},
    {"account": "Cool Amps", "name": "Dana Reyes", "title": "VP Engineering",
     "seniority": "VP", "email": "dana@coolamps.tech",
     "role": "Technical Buyer", "flag": ""},
    {"account": "Hydrolyte Works", "name": "Sam Okoye", "title": "Director of Ops",
     "seniority": "Director", "email": "sam@gmail.com",
     "role": "", "flag": "verify email"},
]

ANGLE = {"Cool Amps": "Argonne ReCell license + CMI Li-ion leach patent"}

GAP = [
    # tier-5 wildcard fragment 'solid.state batter' - the regression case
    {"name": "Northfield Cells", "source": "ARPA-E Summit 2026",
     "description": "solid-state battery pilot line", "prequalified": False},
    # tier-4 wildcard fragment 'grid.scale'
    {"name": "Vantablue Power", "source": "DOE Awardees 2026",
     "description": "grid-scale storage integrator", "prequalified": False},
    # multi-list hit -> should be starred / prioritized
    {"name": "Redcliff Separations", "source": "ARPA-E Summit 2026",
     "description": "rare earth separation pilot plant", "prequalified": False},
    {"name": "Redcliff Separations", "source": "DOE Awardees 2026",
     "description": "rare earth separation pilot plant", "prequalified": False},
    # junk filter should drop this one
    {"name": "Battery Materials Association", "source": "ARPA-E Summit 2026",
     "description": "industry association", "prequalified": False},
    # already in the book -> should be excluded from the gap list
    {"name": "Cool Amps", "source": "DOE Awardees 2026",
     "description": "battery recycling", "prequalified": False},
]


@pytest.fixture(scope="module")
def built(tmp_path_factory):
    """Run the builder once; return (workbook, stdout)."""
    d = tmp_path_factory.mktemp("wf5")
    for fname, payload in [
        ("book.json", BOOK), ("contacts.json", CONTACTS),
        ("angle.json", ANGLE), ("gap_candidates.json", GAP),
    ]:
        (d / fname).write_text(json.dumps(payload), encoding="utf-8")

    out = d / "workbook.xlsx"
    proc = subprocess.run(
        [sys.executable, str(SCRIPT),
         "--book", str(d / "book.json"),
         "--contacts", str(d / "contacts.json"),
         "--angle", str(d / "angle.json"),
         "--gap", str(d / "gap_candidates.json"),
         "--out", str(out)],
        capture_output=True, text=True, cwd=str(d),
    )
    assert proc.returncode == 0, f"builder failed:\nSTDOUT:\n{proc.stdout}\nSTDERR:\n{proc.stderr}"
    assert out.is_file(), "builder reported success but wrote no .xlsx"
    return openpyxl.load_workbook(out), proc.stdout


def test_script_exists():
    assert SCRIPT.is_file(), f"builder missing at {SCRIPT}"


def test_expected_sheets_present(built):
    wb, _ = built
    for name in ["READ ME", "Account Scorecard", "Buying Influences",
                 "Blue Sheet", "Gap Analysis"]:
        assert name in wb.sheetnames, f"missing sheet {name!r} (got {wb.sheetnames})"


def test_read_me_is_first_sheet(built):
    wb, _ = built
    assert wb.sheetnames[0] == "READ ME"


def test_every_book_account_reaches_the_scorecard(built):
    wb, _ = built
    text = "\n".join(
        str(c.value) for row in wb["Account Scorecard"].iter_rows() for c in row
        if c.value is not None
    )
    for acct in ("Cool Amps", "Hydrolyte Works", "Quiet Cases"):
        assert acct in text, f"{acct!r} missing from Account Scorecard"


def test_contacts_land_on_buying_influences(built):
    wb, _ = built
    text = "\n".join(
        str(c.value) for row in wb["Buying Influences"].iter_rows() for c in row
        if c.value is not None
    )
    assert "Lonnie Garris" in text
    assert "Economic Buyer" in text


def _gap_rows(wb):
    """Gap Analysis: header on row 2, data from row 3. Returns {name: row dict}."""
    ws = wb["Gap Analysis"]
    headers = [c.value for c in ws[2]]
    rows = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        d = dict(zip(headers, row))
        rows[str(d["Company"]).lstrip("* ").strip()] = d
    return rows


def test_junk_and_in_book_companies_are_excluded(built):
    wb, _ = built
    rows = _gap_rows(wb)
    assert "Battery Materials Association" not in rows, "junk filter did not drop the association"
    assert "Cool Amps" not in rows, "a company already in the book leaked into Gap Analysis"


def test_multi_list_company_is_flagged(built):
    wb, _ = built
    rows = _gap_rows(wb)
    assert "Redcliff Separations" in rows
    assert rows["Redcliff Separations"]["On lists"] == 2
    assert "PRIORITY" in rows["Redcliff Separations"]["Suggested next step"]


# --- the regression the smoke test exists for -----------------------------
@pytest.mark.parametrize(
    "company,expected_fit",
    [
        ("Northfield Cells", 5),   # 'solid.state batter' - tier 5, not tier-4 'battery'
        ("Vantablue Power", 4),    # 'grid.scale' - tier 4, not tier-3 'storage'
    ],
)
def test_wildcard_keyword_fragments_score_at_the_right_tier(built, company, expected_fit):
    """re.escape'ing the taxonomy fragments demotes these by one tier.

    Vantablue at 3 instead of 4 also silently changes behaviour: a
    pre-qualified list entry force-includes at >=4, so the demotion would
    drop it from the shortlist entirely.
    """
    wb, _ = built
    rows = _gap_rows(wb)
    assert company in rows, f"{company!r} missing from Gap Analysis (got {sorted(rows)})"
    assert rows[company]["Fit (0-5)"] == expected_fit, (
        f"{company!r} scored {rows[company]['Fit (0-5)']}, expected {expected_fit}. "
        "If this fails, the keyword fragments are being escaped again."
    )
